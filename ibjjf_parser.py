# --- Terminal installs ---
# python -m ensurepip
# python -m pip install BeautifulSoup4
# python -m pip install requests
# python -m pip install pandas
# python -m pip install openpyxl
# python -m pip install pipreqs

# --- Libraries ---
from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import Font, PatternFill, Alignment
from sys import exit

# --- HTML Color Code Constants ---
# https://htmlcolorcodes.com/
BLUE = "00CCFF"
PURPLE = "FF00FF"
BROWN = "993300"
BLACK = "000000"
RED = "FF0000"
GREEN_HEADER = "09c97d"
BLUE_URL = "0000ff"

# --- User Inputs ---
team = "G13 BJJ USA"    # The team name is case sensitive.  Some team names have a space " " at the end of the name
tourney_id = "2412"   
club_id = "4440"        # Comes from searching a team from order of fights page
filename = "brackets"

# --- Misc. Variables ---
filename = filename + ".xlsx"
bjjcompsystem = "https://www.bjjcompsystem.com"
gender_category = "?gender_id="
brackets = f"https://www.bjjcompsystem.com/tournaments/{tourney_id}/categories"
registration_url = f"https://www.ibjjfdb.com/ChampionshipResults/{tourney_id}/PublicAcademyRegistration?lang=en-US"
order_of_fights_url = f"https://www.bjjcompsystem.com/tournaments/{tourney_id}/tournament_days/by_club?club_id={club_id}"

# -------------------------------------- Part 1: Generating Initial Excel file --------------------------------------

# Parse IBJJF'S Athlete Registration List
response = requests.get(registration_url)
soup = BeautifulSoup(response.content, 'html.parser')

# Get team's athletes and clean data
athletes = soup.find_all("script")[4].get_text()
# Cut to the left of the string that contains the team's name
athletes = athletes.split(f'{team}",', 1)
athletes.remove(athletes[0])
x = ']},{"'
athletes = athletes[0].split(x, 1)
athletes = athletes[0]
athletes = athletes.replace('"AthleteCategory":[', '')
athletes = athletes.replace('{"FriendlyCategoryName":"', '')
athletes = athletes.replace('","AthleteName":"', ',')
athletes = athletes.replace('"},', ',')
athletes = athletes.replace('"}', '')
athletes = athletes.split(",")

csv_dictionary = {
    'DateTime': [],
    'Time': [],
    'Mat': [],
    'Division': [],
    'Weight Class': [],
    'Name': [],
    'Classification': []
}

rank_list = []

for i in range(0, len(athletes)):
    # Store the rank, age_group, gender, and weight_class
    if i % 2 == 0:
        athlete_info = athletes[i].split("/")
        rank = athlete_info[0]
        age_group = athlete_info[1]
        gender = athlete_info[2]
        weight_class = athlete_info[3]

        # Remove unwanted spaces
        rank = rank[0:-1]
        age_group = age_group[1:-1]
        weight_class = weight_class[1:]

        # Remove weight
        weight_class = weight_class.split(' (')
        weight_class = weight_class[0]

        # Get divsion info, ex. "Master 1" --> "M1"
        if (age_group[0] == "A") or (age_group[0] == "J"):          # "Adult" class and "Juvenile class"
            division = age_group[0]
        else:                                                       # "Master" class
            division = age_group[0] + age_group[-1] 

        # Append to rank_list
        rank_list.append(rank)

        # Mutate weight class name to match brackets
        if weight_class != "Open Class":
            wc = weight_class.replace("-", " ")
        else:
            wc = weight_class

        # Store classification (Example: "M1/F/BLUE/Light") so we can search for it later
        classification = division + "/" + gender[1] + "/" + rank + "/" + wc

        # Append to csv_dictionary
        csv_dictionary['DateTime'].append("TBD")
        csv_dictionary['Time'].append("TBD")
        csv_dictionary['Mat'].append("TBD")
        csv_dictionary['Division'].append(division)
        csv_dictionary['Classification'].append(classification)
        csv_dictionary['Weight Class'].append(weight_class)

    # Store the athlete's name
    else:
        name = athletes[i]
        # Append to team_dictionary 
        csv_dictionary['Name'].append(name)

# Push 'csv_dictionary' to MS Excel (.xlsx) with Pandas  
df = pd.DataFrame.from_dict(csv_dictionary)

# dataframe to csv
df.to_excel(filename, columns=['DateTime', 'Time', 'Mat', 'Division', 'Weight Class', 'Name'], index=False, header=True)

# Format Excel File with OpenPyXL
# Load the already created Excel file
wb = load_workbook(filename)

# Get the active sheet
ws = wb.active

# Adjust column widths
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 30

# Change cell colors for different belt ranks
for i in range(0, len(rank_list)):
    cell = "D" + str(i + 2) # Start at cell "D2"
    if rank_list[i] == "WHITE":
        pass
    elif rank_list[i] == "BLUE":
        ws[cell].fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    elif rank_list[i] == "PURPLE":
        ws[cell].fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
    elif rank_list[i] == "BROWN":
        ws[cell].fill = PatternFill(start_color=BROWN, end_color=BROWN, fill_type="solid")
    elif rank_list[i] == "BLACK":
        ws[cell].font = Font(color=RED) 
        ws[cell].fill = PatternFill(start_color=BLACK, end_color=BLACK, fill_type="solid")

# Change cell colors of column names
ws["A1"].fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")
ws["B1"].fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")
ws["C1"].fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")
ws["D1"].fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")
ws["E1"].fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")
ws["F1"].fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")

# Center align rows 1 to 100 for columns A to E
for row in range(1, 500):
    ws["A" + str(row)].alignment = Alignment(horizontal='center', vertical='center')
    ws["B" + str(row)].alignment = Alignment(horizontal='center', vertical='center')
    ws["C" + str(row)].alignment = Alignment(horizontal='center', vertical='center')
    ws["D" + str(row)].alignment = Alignment(horizontal='center', vertical='center')
    ws["E" + str(row)].alignment = Alignment(horizontal='center', vertical='center')
    ws["F" + str(row)].alignment = Alignment(horizontal='center', vertical='center')

# Save the file and close the connection
wb.save(filename)
wb.close()

print("\nExcel file created.")
# exit()
# Uncomment the "exit()" above if you only want the intial Excel file.  

# -------------------------------------- Part 2: Getting URLs of brackets --------------------------------------
# Scrape age group, gender, rank, & weight class from each bracket
brackets_classification = []
bracket_urls = []

# Iterate through both male (i=1) and female brackets (i=2)
for i in range(1, 3):   
    response = requests.get(brackets + gender_category + str(i))
    brackets_soup = BeautifulSoup(response.content, 'html.parser')
    bracket_count = int(len(brackets_soup.find_all("div", {"class": "category-card__age-division"})) / 2)

    # Get data
    for j in range(0, bracket_count):
        age_group = brackets_soup.find_all("div", {"class": "category-card__age-division"})[j].get_text()
        rank = brackets_soup.find_all('span', class_="category-card__label category-card__belt-label")[j].get_text()
        weight = brackets_soup.find_all('span', class_="category-card__label category-card__weight-label")[j].get_text()
        # Clean data
        age_group = age_group.strip()
        rank = rank.strip() 
        weight = weight.strip()

        # Reformat age group (division).  Get divsion info, ex. "Master 1" --> "M1"
        if (age_group[0] == "A") or (age_group[0] == "J"):          # "Adult" class and "Juvenile class"
            age_group = age_group[0]
        else:                                                       # "Master" class
            age_group = age_group[0] + age_group[-1]  

        if i == 1:
            gender = "M"
        else:
            gender = "F"

        data = age_group + "/" + gender + "/" + rank + "/" + weight
        brackets_classification.append(data)

    # Scrape URL from each bracket
    body = brackets_soup.find('div', attrs={'class' : 'row'})
    for bracket in body.find_all('a', href=True):
        url = bracket['href']
        # Append each url to list
        bracket_urls.append(url)

# Store each url of each competitor for the team of interest
list_of_urls = []
for i in range(0, len(csv_dictionary["Classification"])):
    # Cases where the competitor has a bracket with other competitors
    if csv_dictionary["Classification"][i] in brackets_classification:
        index = brackets_classification.index(csv_dictionary["Classification"][i])
        list_of_urls.append(bjjcompsystem + bracket_urls[index])
    else:
        list_of_urls.append("No bracket")

# Load the already created Excel file
wb = load_workbook(filename)

# Get the active sheet
ws = wb.active

for i in range(0, len(list_of_urls)):
    if (list_of_urls[i] != "No bracket"):
        # Convert each competitor's name into a hyperlink
        ws["F" + str(i + 2)].hyperlink = list_of_urls[i]
        # Style the competitor's name into a hyperlink
        ws["F" + str(i + 2)].font = Font(color=BLUE_URL, underline="single") 

# Save the file and close the connection
wb.save(filename)
wb.close()

print("Excel file updated with individual bracket URLs added.")

# -------------------------------------- Part 3: Extracting time and mat assignment --------------------------------------
# Load the already created Excel file
wb = load_workbook(filename)

# Get the active sheet
ws = wb.active

response = requests.get(order_of_fights_url)
soup = BeautifulSoup(response.content, 'html.parser')

mat_assignment = {
    'Athlete': [],
    'DateTime': [],
    'Time': [],
    'Mat': []
}

# Find the first round matches and byes
for ultag in soup.find_all('ul', {'class': 'list-unstyled tournament-day__matches'}):
    for litag in ultag.find_all('li'):
        # A first round match is occuring
        if litag['class'] == ['match--assigned']:
            for spantag in litag.find_all('span'):
                # Get 'Name'
                if spantag['class'] == ['match-card__competitor-description']:
                    for divtag in spantag.find_all('div'):
                        if divtag['class'] == ['match-card__competitor-name']:
                            if divtag.text in csv_dictionary['Name']:
                                mat_assignment['Athlete'].append(divtag.text) 
                # Get 'DateTime' and 'Time
                elif spantag['class'] == ['search-match-header__when']:
                    time_text = spantag.text
                    # Clean data
                    dateTime = time_text[4:9] + " " + time_text[-8:]
                    time = time_text[-8:]
                    mat_assignment['DateTime'].append(dateTime)
                    mat_assignment['Time'].append(time)
                # Get 'Mat'
                elif spantag['class'] == ['search-match-header__where']:
                    mat_text = spantag.text
                    mat_text = mat_text[6:8]
                    # Remove the ":"
                    mat = mat_text.replace(":", "")
                    mat_assignment['Mat'].append(mat)       
        # A first round bye is occuring
        else:
            # Get 'Name'
            for divtag in litag.find_all('div'):
                if divtag['class'] == ['match-card__competitor-name']:
                    if divtag.text in csv_dictionary['Name']:
                        mat_assignment['Athlete'].append(divtag.text)   
            for spantag in litag.find_all('span'):
                # Get 'DateTime' and 'Time
                if spantag['class'] == ['search-match-header__when']:
                    time_text = spantag.text
                    # Clean data
                    dateTime = time_text[4:9] + " " + time_text[-8:]
                    time = time_text[-8:]
                    mat_assignment['DateTime'].append(dateTime)
                    mat_assignment['Time'].append(time)
                # Get 'Mat'
                elif spantag['class'] == ['search-match-header__where']:
                    mat_text = spantag.text
                    mat_text = mat_text[6:8]
                    # Remove the ":"
                    mat = mat_text.replace(":", "")
                    mat_assignment['Mat'].append(mat)  
    # Don't append duplicates 
    break   

for i in range(0, len(mat_assignment['Athlete'])):
    # Match up indices of mat_assignment and csv_dictionary 
    if mat_assignment['Athlete'][i] in csv_dictionary['Name']:
        key = csv_dictionary['Name'].index(mat_assignment['Athlete'][i])

        # Append 'DateTime', 'Time', and 'Mat'
        ws["A" + str(key + 2)] = mat_assignment['DateTime'][i]
        ws["B" + str(key + 2)] = mat_assignment['Time'][i]
        ws["C" + str(key + 2)] = mat_assignment['Mat'][i]

# Save the file and close the connection
wb.save(filename)
wb.close()

print("Mat assignments have been updated!\n")

# Generate requirements.txt file 
# pipreqs C:\Users\Laptop\Desktop\IBJJF_Parser